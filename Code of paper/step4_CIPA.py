import openpyxl
import matplotlib.pyplot as plt
import os

# 读取Excel文件
file_dir = '../2cost importance performanc analysis/results/'
file_path = file_dir + 'performance_importance_and_cost_for_all_courses.xlsx'  # 替换为您的Excel文件路径
workbook = openpyxl.load_workbook(file_path)

# 读取Performance, Importance, Cost数据
performance_sheet = workbook['performance']
importance_sheet = workbook['importance']
cost_sheet = workbook['cost']

# 获取对象名和属性名
# 获取对象名（第一列，从第二行开始）
object_names = [performance_sheet.cell(row=i, column=1).value for i in range(2, performance_sheet.max_row + 1)]
# 获取属性名（第二行，从第二列开始到最后一列）
attribute_names = [performance_sheet.cell(row=1, column=j).value for j in range(2, performance_sheet.max_column + 1)]

# print(object_names)
# print(attribute_names)
#
scores_list = []
attention_list = []
cost_list = []
for i in range(len(object_names)):
    scores = [float(performance_sheet.cell(row=i+2, column=j).value) for j in range(2, performance_sheet.max_column + 1)]
    attention = [float(importance_sheet.cell(row=i+2, column=j).value) for j in range(2, importance_sheet.max_column + 1)]
    cost = [float(cost_sheet.cell(row=i+2, column=j).value) for j in range(2, cost_sheet.max_column + 1)]
    scores_list.append(scores)
    attention_list.append(attention)
    cost_list.append(cost)


def mean_of_2d_list(matrix):
    # Step 1: Flatten the 2D list into a 1D list
    flattened_list = [element for row in matrix for element in row]
    # Step 2: Calculate the sum of all elements
    total_sum = sum(flattened_list)
    # Step 3: Calculate the number of elements
    num_elements = len(flattened_list)
    # Step 4: Calculate the mean
    mean_value = total_sum / num_elements if num_elements > 0 else 0
    return mean_value

scores_mean = mean_of_2d_list(scores_list)
attention_mean = mean_of_2d_list(attention_list)
cost_mean = mean_of_2d_list(cost_list)
# scores_mean = 0.95

def get_cipa(scores, attention, attributes, mean_score, mean_attention, index, course_names, dir, low_or_high):
    # 创建图形
    plt.figure(figsize=(8, 6))
    # 绘制散点图
    plt.scatter(scores, attention, color='blue', label='Attributes')
    for k, name in enumerate(attributes):
        plt.text(scores[k], attention[k], name, fontsize=10, ha='left', va='bottom')

    # 添加均值线
    plt.axhline(mean_attention, color='red', linestyle='--', label=f'Mean Importance = {mean_attention:.2f}')
    plt.axvline(mean_score, color='green', linestyle='--', label=f'Mean Satisfaction = {mean_score:.2f}')

    # 添加象限标签
    # plt.text(mean_x + 1, mean_y + 1, 'Quadrant I\nKeep up the good work!', fontsize=12, ha='left', va='bottom')
    # plt.text(mean_x - 1, mean_y + 1, 'Quadrant II\nOver-prioritized!', fontsize=12, ha='right', va='bottom')
    # plt.text(mean_x - 1, mean_y - 1, 'Quadrant III\nLow priority!', fontsize=12, ha='right', va='top')
    # plt.text(mean_x + 1, mean_y - 1, 'Quadrant IV\nNeed improvement!', fontsize=12, ha='left', va='top')

    # 添加标题和标签
    if low_or_high == 'low':
        plt.title(f'Cost-Importance-Performance Analysis (CIPA) for {course_names[index]} with low cost')
    else:
        plt.title(f'Cost-Importance-Performance Analysis (CIPA) for {course_names[index]} with high cost')
    plt.xlabel('Performance')
    plt.ylabel('Importance')
    plt.legend()
    plt.grid(True)

    # 调整布局以避免标签重叠（如果需要）
    plt.tight_layout()

    # 保存图形到文件
    if low_or_high == 'low':
        file_name = f'{course_names[index]}_cipa_low_cost.png'
    else:
        file_name = f'{course_names[index]}_cipa_high_cost.png'
    file_path = os.path.join(dir, file_name)
    plt.savefig(file_path)

    # 关闭当前图形以避免内存泄漏（当生成多个图形时很重要）
    plt.close()


for i in range(len(object_names)):
    index_low_cost = [j for j in range(len(scores_list[i])) if cost_list[i][j] <= cost_mean]
    index_high_cost = [j for j in range(len(scores_list[i])) if cost_list[i][j] > cost_mean]
    scores_list_low_cost = [scores_list[i][j] for j in index_low_cost]
    attention_list_low_cost= [attention_list[i][j] for j in index_low_cost]
    attributes_low = [attribute_names[j] for j in index_low_cost]
    scores_list_high_cost = [scores_list[i][j] for j in index_high_cost]
    attention_list_high_cost = [attention_list[i][j] for j in index_high_cost]
    attributes_high = [attribute_names[j] for j in index_high_cost]
    get_cipa(scores_list_low_cost, attention_list_low_cost, attributes_low, scores_mean, attention_mean, i, object_names, file_dir, 'low')
    get_cipa(scores_list_high_cost, attention_list_high_cost, attributes_high, scores_mean, attention_mean, i,
            object_names, file_dir, 'high')



