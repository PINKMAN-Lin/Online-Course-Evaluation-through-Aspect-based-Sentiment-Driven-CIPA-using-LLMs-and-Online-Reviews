import openpyxl
workbook = openpyxl.load_workbook(f"../2cost importance performanc analysis/aspects.xlsx")
sheet = workbook.active
aspects = []
aspects_and_explains = {}
cost_rankings = {}

last_primary_indicator = ''
for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
    # primary_indicator, secondary_indicator, explanation = row[0], row[1], row[2]
    # criteria_and_explains[secondary_indicator] = explanation
    # if primary_indicator:
    #     criteria[primary_indicator] = [secondary_indicator]
    #     last_primary_indicator = primary_indicator
    # else:
    #     criteria[last_primary_indicator].append(secondary_indicator)
    aspects.append(row[1])
    aspects_and_explains[row[1]] = row[2]
    cost_rankings[row[1]] = int(row[3])
# print("Criteria:", criteria)
# print("Criteria and Explains:", criteria_and_explains)




criteria2score = {i:{'正面':0, '负面':0, '成本':0} for i in aspects}

def cost_calculation(rankings):
    denominator = sum([len(rankings) - rankings[i] + 1 for i in rankings.keys()])
    cost = {i:(len(rankings) - rankings[i] + 1)/denominator for i in rankings.keys()}
    return cost

cost = cost_calculation(cost_rankings)

import os
from langchain_openai import ChatOpenAI

from langchain.prompts import PromptTemplate
from langchain.chains import LLMChain
import asyncio

api_key = os.getenv("ZHIPU_API_KEY")
llm = ChatOpenAI(
        model="glm-4-flash",
        openai_api_key=api_key,  # 请填写您自己的APIKey
        openai_api_base="https://open.bigmodel.cn/api/paas/v4/"
    )


import xlrd

dataname = '人工智能5'
wb = xlrd.open_workbook(f"../data_after_preprocessing/{dataname}.xls")
sheet = wb.sheet_by_index(0)  # 通过索引的方式获取到某一个sheet，现在是获取的第一个sheet页，也可以通过sheet的名称进行获取，sheet_by_name('sheet名称')
rows = sheet.nrows  # 获取sheet页的行数，一共有几行
names = []
reviews = []
times = []
likes = []
class_order = []
rankings = []
for i in range(1, rows):
    names.append(str(sheet.cell(i, 0).value).strip())
    reviews.append(str(sheet.cell(i, 1).value).strip())
    times.append(str(sheet.cell(i, 2).value).strip())
    likes.append(str(sheet.cell(i, 3).value).strip())
    class_order.append(str(sheet.cell(i, 4).value).strip())
    rankings.append(str(sheet.cell(i, 5).value).strip())


criteria_for_each_review = [{j:'0' for j in aspects} for _ in range(len(reviews))]

template = '''
    帮我分析课程评论"{review}"中关于方面"{aspect}"的情感，
    其中关于方面“{aspect}”情感的评判标准为“{explain}”。
    如果是肯定的，回复1
    如果是否定的，回复-1
    如果评论与“{aspect}”无关，回复0。即评论中不存在与“{aspect}”的近似表达，回复0。
    请不要过度延伸与解读评论内容，评论内容明确表达了方面“{aspect}”才能回复1或者-1。例如，"非常好，很实用的课程。"这条评论明显相关的方面只有"课程内容的实用性"，在课程内容的实用性方面回复1，其他方面回复都应该是0。
    即评论与“{aspect}”的相关标准是，评论中出现“{aspect}”的近似表达。
    请一步步思考后，再给出最后的回复，务必确保，评论的确提到了“{aspect}”并且关于评论关于它是肯定的态度，才能回复1，评论的确提到了“{aspect}”并且关于评论关于它是否定的态度，才能回复-1。其他一切情况回复0。    
    只回复数字，不要加入你的分析。    
    即最后的回复是1，-1，0三者之一。
    '''


prompt = PromptTemplate.from_template(template)
# print(prompt1)

chain = LLMChain(llm=llm, prompt=prompt)

wb1 = openpyxl.Workbook()
ws1 = wb1.active
ws1.append(['方面', '正面', '负面', '成本'])

wb2 = openpyxl.Workbook()
ws2 = wb2.active
ws2.append(['用户昵称','评论内容', '评论时间', '点赞数','第几次课程', '评分'] + aspects)



async def asy_chain_tool(chain, name, review, time, like, class_ord, ranking, aspects, explains, semaphore):
    async with semaphore:
        input_list = [{'review': review, 'aspect': i, 'explain':explains[i]} for i in aspects]
        resp = await chain.aapply(input_list)
        result = {}
        result['用户昵称'] = name
        result['评论内容'] = review
        result['评论时间'] = time
        result['点赞数'] = like
        result['第几次课程'] = class_ord
        result['评分'] = ranking
        result.update({aspects[i]: resp[i]['text'] for i in range(len(aspects))})
        print(result)
        return result


async def asy_chain():
    semaphore = asyncio.Semaphore(5)
    tasks = [asy_chain_tool(chain, names[i], reviews[i], times[i], likes[i], class_order[i], rankings[i], aspects, aspects_and_explains, semaphore) for i in range(len(reviews))]
    # tasks = [asy_chain_tool(chain, names[i], reviews[i], times[i], likes[i], class_order[i], rankings[i], aspects,
    #                         aspects_and_explains, semaphore) for i in range(100)]
    results_all = await asyncio.gather(*tasks)
    return results_all


results_all = asyncio.run(asy_chain())
print(results_all[:10])


for i in range(len(reviews)):
    print(i)
    for j in criteria2score.keys():
        if results_all[i][j] == '1':
            criteria2score[j]['正面'] += 1
            criteria_for_each_review[i][j] = '1'
        elif results_all[i][j] == '-1':
            criteria2score[j]['负面'] += 1
            criteria_for_each_review[i][j] = '-1'
    ws2.append([results_all[i]['用户昵称'], results_all[i]['评论内容'], results_all[i]['评论时间'], results_all[i]['点赞数'], results_all[i]['第几次课程'],results_all[i]['评分']] + [results_all[i][j] for j in aspects])
      # else:
      #   criteria_for_each_review[i][j][k] = '2'
wb2.save(f"../2cost importance performanc analysis/results/{dataname}_attributes_for_each_review.xlsx")

for i in criteria2score.keys():
    ws1.append([i, criteria2score[i]['正面'], criteria2score[i]['负面'], cost[i]])
wb1.save(
    f"../2cost importance performanc analysis/results/{dataname}_attributes_scores.xlsx")

