import openpyxl

wb_w = openpyxl.Workbook()
ws1 = wb_w.active
ws1.title = "performance"
ws2 = wb_w.create_sheet(title="importance")
ws3 = wb_w.create_sheet(title="cost")


# scores_all = []
# attention_all = []
aspects = []
for i in range(1, 6):
    dataname = f'人工智能{i}'
    wb_r = openpyxl.load_workbook(
        f"../2cost importance performanc analysis/results/{dataname}_attributes_scores.xlsx")
    ws_r = wb_r.active
    positive_list = []
    negative_list = []
    cost_list = []
    for row_idx, row in enumerate(ws_r.iter_rows(min_row=2, values_only=True), start=2):
        if i == 1:
            aspects.append(row[0])
        positive_list.append(int(row[1]))
        negative_list.append(int(row[2]))
        cost_list.append(row[3])
    print(positive_list)
    print(negative_list)
    scores = [positive_list[k]/(positive_list[k]+negative_list[k] + 0.01) for k in range(len(positive_list))]
    attention = [(positive_list[k]+negative_list[k])/sum(positive_list + negative_list) for k in range(len(positive_list))]
    if i == 1:
        ws1.append(['Courses'] + aspects)
        ws2.append(['Courses'] + aspects)
        ws3.append(['Courses'] + aspects)

    ws1.append([dataname] + [str(k) for k in scores])
    ws2.append([dataname] + [str(k) for k in attention])
    ws3.append([dataname] + [k for k in cost_list])

wb_w.save(f"../2cost importance performanc analysis/results/performance_importance_and_cost_for_all_courses.xlsx")